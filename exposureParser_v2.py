import os
import sys
import argparse
import openpyxl
from datetime import datetime
import re

parser = argparse.ArgumentParser(description='Parses all exposure .xlsx files in a directory and outputs to a file.')
parser.add_argument('-i', '--inputdir', help='Directory to search. If at least one directory is not specified\
                        the program prints help.', required=True)

# if no arguments are provided to the program print help
if len(sys.argv) == 1:
    parser.print_help()
    sys.exit(1)
args = parser.parse_args()

# static tuples (headers and cell locations for hospitals and facilities)
# removed General Liability exposure data for facilities per Jhavid 12/12/2019
HOSPITAL_HEADERS = ('Policy Holder','Policy Number','YearFolder','PolicyTermStart','PolicyTermEnd','Location','Beds-AcuteCare','Beds-Extended','Beds-Skilled','Beds-Personal','Visits-InpatientSurgeries','Visits-Births','Visits-OutpatientSurgeries','Visits-ER','Visits-OtherOPVs','Visits-HomeHealth','Visits-PhysicalTherapy','Visits-MentalHealth','Visits-SubstanceAbuse','Visits-UrgiCenter','Visits-DialysisCenter','Receipts-DurableMedEquip','Receipts-XRayImaging','Receipts-Pharmacy','Equivalencies-CRNAs','Equivalencies-EMTs','GeneralLiability-HospitalGLPrem','GeneralLiability-ApartmentExposure','GeneralLiability-DayCareExposure','GeneralLiability-DwellingsExposure','GeneralLiability-FitnessCenter','GeneralLiability-OCBExposure','GeneralLiability-ParkingExposure','GeneralLiability-StorageExposure','GeneralLiability-VacantLandExposure','GeneralLiability-EmployeeBenefits','Filename')
HOSPITAL_CELLS_2010 = ("B22","C22","D22","E22","F22","G22","H22","I22","J22","K22","B31","C31","D31","E31","F31","G31","H31","I31","J31","K31","B44","C44","D44","E44","F44","G44","H44","I44","J44","K44")
HOSPITAL_CELLS_2006 = ('B21','C21','D21','E21','F21','G21','H21','I21','J21','K21','B30','C30','D30','E30','F30','G30','H30','I30','J30','K30','B43','C43','D43','E43','F43','G43','H43','I43','J43','K43')
HOSPITAL_CELLS_2005 = ('B22','C22','D22','E22','F22','G22','H22','I22','J22','K22','B31','C31','D31','E31','F31','G31','H31','I31','J31','K31','B43','C43','D43','E43','F43','G43','H43','I43','J43','K43')
FACILITY_HEADERS_2015 = ('Policy Holder','Policy Number','YearFolder','PolicyTermStart','PolicyTermEnd','Location','Donations-BloodBank','Donations-OrganBankNoProcessing','Donations-OrganBankProcessing','Staff-AmbulanceService','Staff-MedicalRegistry','Staff-ParamedicEMT','Staff-CRNA','Receipts-DentalLaboratory','Receipts-HomeCareDurableEquip','Receipts-LaboratoryAllOther','Receipts-MedicalLaboratory','Receipts-OpticalEstablishment','Receipts-OcularLaboratory','Receipts-PathologyLaboratory','Receipts-Pharmacy','Receipts-XRayImagingCenter','Visits-AbortionClinics','Visits-BirthingCenter','Visits-CardiacRehabilitation','Visits-CollegeHealthCenter','Visits-CommunityHealthCenter','Visits-CrisisCenter','Visits-DevelopmentalDisability','Visits-DialysisCenter','Visits-Emergicenter','Visits-EndoscopyCenter','Visits-HomeCarePersonal','Visits-HomeCareSkilled','Visits-HomeCareRehab','Visits-HomeCareIntravenous','Visits-HomeCareRespiratory','Visits-HospiceCare','Visits-LithotripsyServices','Visits-Medispa','Visits-MentalHealthCounseling','Visits-MunicipalHealthDepartment','Visits-OncologyServices','Visits-PhyOccupationalRehab','Visits-RetailConvenienceClinic','Visits-SubstanceAbuseCoun','Visits-SurgiCenter','Visits-TraumaRehabilitation','Visits-TraumaRehabilitationTherapy','Visits-TraumaRehabilitationAllOther','Visits-Urgicenter','Visits-WeightLossCenter','School-ChiropracticSchool','School-CRNASchool','School-DentalSchool','School-EMTSchool','School-MedicalSchool','School-NursingSchool','School-OptometrySchool','School-AllOtherSchool','Beds-BirthingCenter','Beds-CrisisCenter','Beds-CardiovascularRehabilitation','Beds-DevelopmentalDisability','Beds-HospiceBeds','Beds-LongTermAcuteCare-FP','Beds-LongTermAcuteCare-NP','Beds-LongTermSkilled-FP','Beds-LongTermSkilled-NP','Beds-LongTermIntermediate-FP','Beds-LongTermIntermediate-NP','Beds-LongTermAssistedLiving-FP','Beds-LongTermAssistedLiving-NP','Beds-LongTermPersonalCare-FP','Beds-LongTermPersonalCare-NP','Beds-MentalHealthCounseling','Beds-PhysicalOccupational','Beds-SleepLaboratory','Beds-SubstanceAbuseCounseling','Beds-SurgiCenter','Beds-TraumaRehab','Beds-TraumaRehabTransitionalLiving','Filename')
FACILITY_HEADERS_2004 = ('Policy Holder','Policy Number','YearFolder','PolicyTermStart','PolicyTermEnd','Location','Receipts-OpticalEstablishment','Receipts-OpticalAllOtherEstab','Receipts-OcularEstablishment','Receipts-DentalLaboratory','Receipts-XRayImagingCenter','Receipts-PathologyLaboratory','Receipts-MedicalLaboratory','Receipts-Pharmacy','Receipts-HomeCareDurableEquip','Visits-CommunityHealthCenter','Visits-CollegeHealthCenter','Visits-Urgicenter','Visits-SubstanceAbuseCoun','Visits-Emergicenter','Visits-PhyOccupRehab','Visits-CardiacRehab','Visits-Surgicenter','Visits-DialysisCenter','Visits-MentalHealthCounseling','Visits-DevelopmentalDisability','Visits-SubstanceAbuseSkilledMedical','Visits-HospiceCare','Visits-HomeCarePersonal','Visits-HomeCareSkilled','Visits-HomeCareRehab','Visits-HomeCareIntravenous','Visits-HomeCareRespiratory','Visits-DialysisCenter','Visits-AmbulanceService','StudentFacility-SchoolChiropracticTeacher','StudentFacility-StudentFacility-SchoolChiropracticStudent','StudentFacility-SchoolMedicalTeacher','StudentFacility-SchoolMedicalStudent','StudentFacility-SchoolDentalTeacher','StudentFacility-SchoolDentalStudent','StudentFacility-SchoolOptometryTeacher','StudentFacility-SchoolOptometryStudent','StudentFacility-SchoolNursingTeacher','StudentFacility-SchoolNursingStudent','Filename')
FACILITY_CELLS_2015 = ('B19','C19','D19','E19','F19','G19','H19','B28','C28','D28','E28','F28','G28','H28','I28','J28','B37','C37','D37','E37','F37','G37','H37','I37','J37','K37','B46','C46','D46','E46','F46','G46','H46','I46','J46','K46','B55','C55','D55','E55','F55','G55','H55','I55','J55','K55','B64','C64','D64','E64','F64','G64','H64','I64','B73','C73','D73','E73','F73','G73','H73','I73','J73','K73','B82','C82','D82','E82','F82','G82','H82','I82','J82','K82','B91','C91')
FACILITY_CELLS_2011 = ('B19','C19','D19','E19','F19','G19','H19','B28','C28','D28','E28','F28','G28','H28','I28','J28','B37','C37','D37','E37','F37','G37','H37','I37','J37','K37','B45','C45','D45','E45','F45','G45','H45','I45','J45','K45','B53','C53','D53','E53','F53','G53','H53','I53','J53','K53','B61','C61','D61','E61','F61','G61','H61','I61','B69','C69','D69','E69','F69','G69','H69','I69','J69','K69','B77','C77','D77','E77','F77','G77','H77','I77','J77','K77','B85','C85')
FACILITY_CELLS_2004 = ('B18','C18','D18','E18','F18','G18','H18','I18','J18','B27','C27','D27','E27','F27','G27','H27','I27','J27','K27','B35','C35','D35','E35','F35','G35','H35','I35','J35','K35','B43','C43','D43','E43','F43','G43','H43','I43','J43','K43')

# create output lists to hold lists of exposures for the workbooks
hospitalExposureList = []
hospitalExposureList.append(HOSPITAL_HEADERS)
facilityExposureList = []
facilityExposureList.append(FACILITY_HEADERS_2015)
facilityExposureList2004 = []
facilityExposureList2004.append(FACILITY_HEADERS_2004)

# create output and error log file in current working directory
logfile = os.path.join(os.getcwd(), 'output_' + datetime.now().strftime("%Y%m%d-%H%M%S") + '.log')
errfile = os.path.join(os.getcwd(), 'error_' + datetime.now().strftime("%Y%m%d-%H%M%S") + '.log')

ob = openpyxl.Workbook()
out_filename = 'c:/projects/exposureparser/ExposureBook.xlsx'
hs1 = ob.active
hs1.title = "HospitalExposures"
fs1 = ob.create_sheet("Sheet")
fs1.title = "FacilitiesExposures Format1"
fs2 = ob.create_sheet("Sheet2")
fs2.title = "FacilitiesExposures Format2"

try:
    d = open(logfile, 'w+')
    err = open(errfile, 'w+')
    print ("Parsing exposure .xlsx files in " + args.inputdir + " and subdirectories.\n")
    d.write ("Parsing exposure .xlsx files in " + args.inputdir + " and subdirectories.\n")

    for dirName, subdirList, fileList in os.walk(args.inputdir):
        print ("Searching for and parsing exposure .xlsx files in " + dirName + "\n")
        d.write ("Searching for and parsing exposure .xlsx files in " + dirName + "\n")
        for workfile in fileList:
            # make sure we're opening a .xlsx file
            if workfile.endswith('.xlsx'):
                try:
                    # filepath = os.path.join(args.inputdir, workfile)
                    filepath = os.path.join (dirName, workfile)

                    # try open xlsx file with openpyxl and throw exception if failure
                    try:
                        print ("Trying to open " + filepath + "\n")
                        d.write ("Trying to open " + filepath + "\n")
                        wb = openpyxl.load_workbook(filename=filepath, data_only=True)

                        # zero out variables
                        locationList = []
                        isHospital = False
                        isFacility = False
                        alternativeDateField = False

                        # create location list, parse though sheetList to get all Location sheetnames (except all locations) and add to locationList
                        # if no locations then figure out what sheet to get data from
                        sheetNames = wb.sheetnames
                        for i in sheetNames:
                            if i.startswith("Location"):
                                locationList.append(i)
                        # if locationList has stuff in it inputSheet should be input page
                        if len(locationList) > 0:
                            inputSheet = wb["Input Page"]
                            # since we have an input page and locations we need to figure out if this is a facility or hospital
                            facilityString = inputSheet["A15"].value
                            facilityString2 = inputSheet["A1"].value
                            if facilityString != None and 'facility' in facilityString.lower() or facilityString2 != None and 'facility' in facilityString2.lower():
                                isFacility = True
                            else:
                                isHospital = True
                        
                    # if location list is empty
                        if not locationList:
                            alternativeDateField = True
                            if 'Hospital' in sheetNames:
                                locationList.append("Hospital")
                                inputSheet = wb["Hospital"]
                                isHospital = True
                            elif 'Primary & Excess Worksheet' in sheetNames:
                                locationList.append("Primary & Excess Worksheet")
                                inputSheet = wb["Primary & Excess Worksheet"]
                                isHospital = True
                            elif 'Primary Worksheet' in sheetNames:
                                locationList.append("Primary Worksheet")
                                inputSheet = wb["Primary Worksheet"]
                                isHospital = True
                            elif 'Rating Worksheet' in sheetNames:
                                locationList.append("Rating Worksheet")
                                inputSheet = wb["Rating Worksheet"]
                                isFacility = True
                            elif 'Master' in sheetNames:
                                locationList.append("Master")
                                inputSheet = wb["Master"]
                                isFacility = True
                            else:
                                # if none of the above are true we have no idea what kind of workbook this is
                                # close the workbook and move on to next file
                                print ('Unable to identify workbook type for ' + filepath + '. Skipping.\n')
                                err.write ('Unable to identify workbook type for ' + filepath + '. Skipping.\n')
                                wb.close()
                                continue

                        # use regex to try grab policy number out of the filename just in case it's blank inside the input page
                        possiblePolicyNum = re.findall('[a-zA-Z]{3}[0-9]{3,6}', workfile)
                        # read policy number, start and end dates (fields are same on both types of sheets)
                        policyHolder = inputSheet["C3"].value
                        if policyHolder == None:
                            policyHolder = 'None'
                        policyNumber = inputSheet["C7"].value
                        # use regex to get policy number (if present) from filename in case policy number from cell C7 is blank
                        if policyNumber == None:
                            if possiblePolicyNum:
                                policyNumber = possiblePolicyNum[0]
                            else:
                                policyNumber = 'None'
                        # use regex to try grab year from the folder structure in case someone didn't put policy dates
                        yearList = re.findall('[0-9]{4}', dirName)
                        if yearList:
                            yearFolder = yearList[0]
                        else:
                            yearFolder = 'None'
                        
                        # if it's an older sheet there's one date field for policies instead of two
                        if alternativeDateField:
                            dates = inputSheet["C6"].value
                            if dates == None:
                                StartDate = 'None'
                                EndDate = 'None'
                            else:
                                # if date is xx/xx/xx to xx/xx/xx we'll try split the string to grab start and end date
                                dateList = dates.split()
                                if len(dateList) == 3:
                                    StartDate = dateList[0]
                                    EndDate = dateList[2]
                                # otherwise put whatever is in the date field in both fields in output
                                else:
                                    StartDate = dateList[0]
                                    EndDate = dateList[0]
                        # we have two date files, so grab start and end date
                        else:
                            StartDate = inputSheet["C5"].value
                            EndDate = inputSheet["C6"].value
                        # if start date is empty put none
                        if StartDate == None:
                            policyStartDate = 'None'
                        else:
                            # if start date isn't a string convert it from a datetime to a string
                            if not isinstance(StartDate, str):
                                policyStartDate = StartDate.strftime("%m/%d/%Y")
                            else:
                                policyStartDate = StartDate
                        # if end date is empty put none
                        if EndDate == None:
                            policyEndDate = 'None'
                        else:
                            # if end date isn't a string convert it from a datetime to a string
                            if not isinstance(EndDate, str):
                                policyEndDate = EndDate.strftime("%m/%d/%Y")
                            else:
                                policyEndDate = EndDate
                        
                        # if the workbook has locations we need to switch the inputsheet to the first location sheet for row validation
                        if not alternativeDateField:
                            inputSheet = wb[locationList[0]]
                        # since rows with exposures change all the time we need to figure out which combination we need to grab
                        # figure out cell combination to use for facility workbook
                        if isFacility:
                            a43 = inputSheet["A43"].value
                            if a43 == None:
                                a43 = 'None'
                            a85 = inputSheet["A85"].value
                            if a85 == None:
                                a85 = 'None'
                            a91 = inputSheet["A91"].value
                            if a91 == None:
                                a91 = 'None'
                            if 'exposure' in a91.lower():
                                CELLS = FACILITY_CELLS_2015
                                OUTLIST = facilityExposureList
                            elif 'exposure' in a85.lower():
                                CELLS = FACILITY_CELLS_2011
                                OUTLIST = facilityExposureList
                            elif 'exposure' in a43.lower():
                                CELLS = FACILITY_CELLS_2004
                                OUTLIST = facilityExposureList2004
                            # if all the above conditions fail we don't know what cells to read so close the workbook and move on
                            else:
                                print ('Workbook looks like a facility but unable to identify exposure rows in ' + filepath + '. Skipping.\n')
                                err.write ('Workbook looks like a facility but unable to identify exposure rows in ' + filepath + '. Skipping.\n')
                                wb.close()
                                continue

                        # figure out cell combination to use for hospital workbook   
                        if isHospital:
                            OUTLIST = hospitalExposureList
                            a21 = inputSheet["A21"].value
                            if a21 == None:
                                a21 = 'None'
                            a22 = inputSheet["A22"].value
                            if a22 == None:
                                a22 = 'None'
                            a30 = inputSheet["A30"].value
                            if a30 == None:
                                a30 = 'None'
                            a31 = inputSheet["A31"].value
                            if a31 == None:
                                a31 = 'None'
                            a43 = inputSheet["A43"].value
                            if a43 == None:
                                a43 = 'None'
                            a44 = inputSheet["A44"].value
                            if a44 == None:
                                a44 = 'None'
                            if 'exposure' in a22.lower() and 'exposure' in a31.lower() and 'exposure' in a44.lower():
                                CELLS = HOSPITAL_CELLS_2010
                            elif 'exposure' in a21.lower() and 'exposure' in a30.lower() and 'exposure' in a43.lower():
                                CELLS = HOSPITAL_CELLS_2006
                            elif 'exposure' in a22.lower() and 'exposure' in a31.lower() and 'exposure' in a43.lower():
                                CELLS = HOSPITAL_CELLS_2005
                            # if all the above conditions fail we don't know what cells to read so close the workbook and move on
                            else:
                                print ('Workbook looks like a hospital but unable to identify exposure rows in ' + filepath + '. Skipping.\n')
                                err.write ('Workbook looks like a hospital but unable to identify exposure rows in ' + filepath + '. Skipping.\n')
                                wb.close()
                                continue
                        
                        # iterate through each location sheet
                        for location in locationList:
                            # make a temp list to hold the location exposure data. we will add to exposure output
                            tempList = []
                            # add policy number and dates for each location along with location number (stripping number off the end of location sheet name) to tempList
                            tempList.extend((policyHolder, policyNumber, yearFolder, policyStartDate, policyEndDate, location))
                            # start grabbing exposures from current location sheet and add to tempList
                            s = wb[location]
                            for i in CELLS:
                                tempList.append(s[i].value)
                            tempList.append(filepath)
                            OUTLIST.append(tempList)
                        
                        print ("File " + workfile + " processed successfully\n")
                        d.write ("File " + workfile + " processed successfully\n")
                        # close current workbook
                        wb.close()
                    except:
                        print ("Could not open " + filepath + " using openpyxl.load_workbook. Check that file is valid.\n")
                        err.write ("Could not open " + filepath + " using openpyxl.load_workbook. Check that file is valid.\n")
                    
                # skip any file we can't open due to permissions issues, file locks, etc
                except IOError:
                    print ("Error opening " + workfile + ". Continuing.\n")
                    err.write ("Error opening " + workfile + ". Continuing.\n")
    
    # write out exposures to workbook
    for row in hospitalExposureList:
        try:
            hs1.append(row)
        except:
            err.write("cannot write row to hospital sheet" + row + '\n')
    for row in facilityExposureList:
        try:
            fs1.append(row)
        except:
            err.write("cannot write row to facility format1 sheet" + row + '\n')
    for row in facilityExposureList2004:
        try:
            fs2.append(row)
        except:
            err.write("cannot write row to facility format2 sheet" + row + '\n')
    

    # save exposure workbook we've written
    ob.save(filename=out_filename)

#  can't open output file due to permissions issues, file locks, etc
except IOError:
    print ('Error opening log file ', logfile, ".\n")