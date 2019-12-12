import os
import sys
import argparse
import openpyxl
from datetime import datetime

parser = argparse.ArgumentParser(description='Parses all exposure .xlsx files in a directory and outputs to a file.')
parser.add_argument('-i', '--inputdir', help='Directory to search. If at least one directory is not specified\
                        the program prints help.', required=True)

# if no arguments are provided to the program print help
if len(sys.argv) == 1:
    parser.print_help()
    sys.exit(1)
args = parser.parse_args()

# static tuples (headers and cell locations for hospitals and facilities)
HOSPITAL_HEADERS = ('Policy Number','PolicyTermStart','PolicyTermEnd','Location','Beds-Acute Care','Beds-Extended','Beds-Skilled','Beds-Personal','Visits-InpatientSurgeries','Visits-Births','Visits-OutpatientSurgeries','Visits-ER','Visits-OtherOPVs','Visits-HomeHealth','Visits-PhysicalTherapy','Visits-MentalHealth','Visits-SubstanceAbuse','Visits-UrgiCenter','Visits-DialysisCenter','Receipts-DurableMedEquip','Receipts-XRayImaging','Receipts-Pharmacy','Equivalencies-CRNAs','Equivalencies-EMTs','GeneralLiability-HospitalGLPrem','GeneralLiability-ApartmentExposure','GeneralLiability-DayCareExposure','GeneralLiability-DwellingsExposure','GeneralLiability-FitnessCenter','GeneralLiability-OCPExposure','GeneralLiability-ParkingExposure','GeneralLiability-StorageExposure','GeneralLiability-VacantLandExposure','GeneralLiability-EmployeeBenefits','Filename')
HOSPITAL_CELLS = ("B22","C22","D22","E22","F22","G22","H22","I22","J22","K22","B31","C31","D31","E31","F31","G31","H31","I31","J31","K31","B44","C44","D44","E44","F44","G44","H44","I44","J44","K44")
FACILITY_HEADERS = ('Policy Number','PolicyTermStart','PolicyTermEnd','Location','Donations-BloodBank','Donations-OrganBankNoProcessing','Donations-OrganBankProcessing','Staff-AmbulanceService','Staff-MedicalRegistry','Staff-ParamedicEMT','Staff-CRNA','Receipts-DentalLaboratory','Receipts-HomeCareDurableEquip','Receipts-LaboratoryAllOther','Receipts-MedicalLaboratory','Receipts-OpticalEstablishment','Receipts-OcularLaboratory','Receipts-PathologyLaboratory','Receipts-Pharmacy','Receipts-XRayImagingCenter','Visits-AbortionClinics','Visits-BirthingCenter','Visits-CardiacRehabilitation','Visits-CollegeHealthCenter','Visits-CommunityHealthCenter','Visits-CrisisCenter','Visits-DevelopmentalDisability','Visits-DialysisCenter','Visits-Emergicenter','Visits-EndoscopyCenter','Visits-HomeCarePersonal','Visits-HomeCareSkilled','Visits-HomeCareRehab','Visits-HomeCareIntravenous','Visits-HomeCareRespiratory','Visits-HospiceCare','Visits-LithotripsyServices','Visits-Medispa','Visits-MentalHealthCounseling','Visits-MunicipalHealthDepartment','Visits-OncologyServices','Visits-PhyOccupationalRehab','Visits-RetailConvenienceClinic','Visits-SubstanceAbuseCoun','Visits-SurgiCenter','Visits-TraumaRehabilitation','Visits-TraumaRehabilitationTherapy','Visits-TraumaRehabilitationAllOther','Visits-Urgicenter','Visits-WeightLossCenter','School-ChiropracticSchool','School-CRNASchool','School-DentalSchool','School-EMTSchool','School-MedicalSchool','School-NursingSchool','School-OptometrySchool','School-AllOtherSchool','Beds-BirthingCenter','Beds-CrisisCenter','Beds-CardiovascularRehabilitation','Beds-DevelopmentalDisability','Beds-HospiceBeds','Beds-LongTermAcuteCare-FP','Beds-LongTermAcuteCare-NP','Beds-LongTermSkilled-FP','Beds-LongTermSkilled-NP','Beds-LongTermIntermediate-FP','Beds-LongTermIntermediate-NP','Beds-LongTermAssistedLiving-FP','Beds-LongTermAssistedLiving-NP','Beds-LongTermPersonalCare-FP','Beds-LongTermPersonalCare-NP','Beds-MentalHealthCounseling','Beds-PhysicalOccupational','Beds-SleepLaboratory','Beds-SubstanceAbuseCounseling','Beds-SurgiCenter','Beds-TraumaRehab','Beds-TraumaRehabTransitionalLiving','GeneralLiability-FacilitySquareFeet','GeneralLiability-StorageExposure','GeneralLiability-DayCareExposure','GeneralLiability-ParkingExposure','GeneralLiability-VacantLandExposure','GeneralLiability-DwellingsExposure','GeneralLiability-ApartmentExposure','GeneralLiability-OCPExposure','GeneralLiability-EmployersLiability','GeneralLiability-EmployeeBenefits','GeneralLiability-FitnessCenter','Filename')
FACILITY_CELLS = ('B19','C19','D19','E19','F19','G19','H19','B28','C28','D28','E28','F28','G28','H28','I28','J28','B37','C37','D37','E37','F37','G37','H37','I37','J37','K37','B46','C46','D46','E46','F46','G46','H46','I46','J46','K46','B55','C55','D55','E55','F55','G55','H55','I55','J55','K55','B64','C64','D64','E64','F64','G64','H64','I64','B73','C73','D73','E73','F73','G73','H73','I73','J73','K73','B82','C82','D82','E82','F82','G82','H82','I82','J82','K82','B91','C91','B109','C109','D109','E109','F109','G109','H109','I109','J109','K109','L109')

# create output lists to hold lists of exposures for the workbooks
hospitalExposureList = []
hospitalExposureList.append(HOSPITAL_HEADERS)
facilityExposureList = []
facilityExposureList.append(FACILITY_HEADERS)

# create output and error log file in current working directory
logfile = os.path.join(os.getcwd(), 'output_' + datetime.now().strftime("%Y%m%d-%H%M%S") + '.log')
errfile = os.path.join(os.getcwd(), 'error_' + datetime.now().strftime("%Y%m%d-%H%M%S") + '.log')

ob = openpyxl.Workbook()
out_filename = 'c:/myprojects/jrproj/ExposureBook.xlsx'
hs1 = ob.active
hs1.title = "HospitalExposures"
fs1 = ob.create_sheet("Sheet")
fs1.title = "FacilitiesExposures"

try:
    d = open(logfile, 'w+')
    err = open(errfile, 'w+')
    print ("Parsing exposure .xlsx files in " + args.inputdir + " and subdirectories.\n")
    d.write ("Parsing exposure .xlsx files in " + args.inputdir + " and subdirectories.\n")

    for dirName, subdirList, fileList in os.walk(args.inputdir):
        print ("Searching for and parsing exposure .xlsx files in " + dirName + "\n")
        d.write ("Searching for and parsing exposure .xlsx files in " + dirName + "\n")
        for workfile in fileList:
            # make sure we're opening a .xlsx file
            if workfile.endswith('.xlsx'):
                try:
                    # filepath = os.path.join(args.inputdir, workfile)
                    filepath = os.path.join (dirName, workfile)

                    # try open xlsx file with openpyxl and throw exception if failure
                    try:
                        print ("Trying to open " + filepath + "\n")
                        d.write ("Trying to open " + filepath + "\n")
                        wb = openpyxl.load_workbook(filename=filepath, data_only=True)

                        # create location list, parse though sheetList to get all Location sheetnames (except all locations) and add to locationList
                        locationList = []
                        for i in wb.sheetnames:
                            if i.startswith('Location'):
                                locationList.append(i)
                        
                        # read policy number, start and end dates (fields are same on both types of sheets)
                        inputSheet = wb['Input Page']
                        policyNumber = inputSheet["C7"].value
                        StartDate = inputSheet["C5"].value
                        if StartDate == None:
                            policyStartDate = 'None'
                        else:
                            policyStartDate = StartDate.strftime("%m/%d/%Y")
                        EndDate = inputSheet["C6"].value
                        if EndDate == None:
                            policyEndDate = 'None'
                        else:
                            policyEndDate = EndDate.strftime("%m/%d/%Y")
                        # identify if sheet is hospital or facility exposure sheet so we can grab correct fields
                        facilityString = inputSheet["A15"].value
                        
                        if facilityString != None and 'FACILITY' in facilityString:
                            CELLS = FACILITY_CELLS
                            OUTLIST = facilityExposureList
                        else:
                            CELLS = HOSPITAL_CELLS
                            OUTLIST = hospitalExposureList
                        
                        # iterate through each location sheet
                        for location in locationList:
                            # make a temp list to hold the location exposure data. we will add to exposure output
                            tempList = []
                            # add policy number and dates for each location along with location number (stripping number off the end of location sheet name) to tempList
                            tempList.extend((policyNumber, policyStartDate, policyEndDate, location[8:]))
                            # start grabbing exposures from current location sheet and add to tempList
                            s = wb[location]
                            for i in CELLS:
                                tempList.append(s[i].value)
                            tempList.append(filepath)
                            OUTLIST.append(tempList)
                        
                        print ("File " + workfile + " processed successfully\n")
                        d.write ("File " + workfile + " processed successfully\n")
                        # close current workbook
                        wb.close()
                    except:
                        print ("Could not open " + filepath + " using openpyxl.load_workbook. Check that file is valid.\n")
                        err.write ("Could not open " + filepath + " using openpyxl.load_workbook. Check that file is valid.\n")
                    
                # skip any file we can't open due to permissions issues, file locks, etc
                except IOError:
                    print ("Error opening " + workfile + ". Continuing.\n")
                    err.write ("Error opening " + workfile + ". Continuing.\n")
    
    # write out exposures to workbook
    for row in hospitalExposureList:
        hs1.append(row)
    for row in facilityExposureList:
        fs1.append(row)
    # save exposure workbook we've written
    ob.save(filename=out_filename)

#  can't open output file due to permissions issues, file locks, etc
except IOError:
    print ('Error opening log file ', logfile, ".")